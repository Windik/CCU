import openpyxl
import sqlite3
# from openpyxl import load_workbook

def dateChanger(source_string):
	new_string = source_string
	if "datetime.datetime(" in new_string:
		while "datetime.datetime(" in new_string:
			# Находим дату и разбиваем на составляющие
			lol = new_string[(new_string.find("datetime.datetime(")+len("datetime.datetime(")):new_string.find(")")].split(', ') # "datetime.datetime(".__sizeof__()
			# Формируем новую дату
			new_date_format = lol[2] + "-" + lol[1] + "-" + lol[0]
			# Модифицируем строку
			new_string = new_string[:new_string.find("datetime.datetime(")] + new_date_format + new_string[new_string.find("),")+1:]
	print(new_string)
		


# открываем файл excel
wb = openpyxl.load_workbook(filename = 'contr.xlsx')

#sheet_ranges = wb['range names']

# получаем имена страниц в виде списка
pages = wb.sheetnames
# print(pages)

# Worksheet
one_page = wb[pages[0]]
temp_string_index = ''
# Перебор строк
'''for i in range(2,19):
	temp_string_index = 'D' + str(i)
	print(one_page[temp_string_index].value)
'''

temp_list = []
full_dict = dict()
for row in range(2,6):
	for i in 'CDEFGJP':
		temp_string_index = i + str(row)
		temp_list.append(one_page[temp_string_index].value)
	full_dict[row] = temp_list
	temp_list = []

for i in range(2,30):
	pass
	#print(full_dict[i])
		# print(temp_string_index, one_page[temp_string_index].value)

#print(str(full_dict))

test_file = open('data.txt','w')
test_file.write(str(full_dict))
test_file.close()

connect = sqlite3.connect("YourDataBase.dbl")
cursor = connect.cursor()

data = ""
list_data = list()
for i in full_dict:
	data =  (str(full_dict[i]))
	list_data = data.strip('[]')

	# Вырезаем часть с датой, разбиваем ее по запятой
	lol = list_data[(list_data.find("datetime.datetime(")+len("datetime.datetime(")):list_data.find(")")].split(', ') # "datetime.datetime(".__sizeof__()
	# 
	
	dateChanger(list_data)
	#list_data = list_data.replace("datetime.datetime(", "")
	#list_data = list_data.replace(")","")
	#while '(' in list_data:
	#	list_data = list_data # [list_data.find('(')+1:list_data.find(')')] + list_data[list_data.find(')')+1:-1]
	#print(list_data)
	print()


for i in range(2,5):

	#query_str = """INSERT INTO organisations VALUES (NULL, """ + data + """)"""
	# print(self.query_str)
	#cursor.execute(query_str)
	pass
#connect.commit()

#for i in range(0,19):
#	print(page['A'+str(i)].value)


